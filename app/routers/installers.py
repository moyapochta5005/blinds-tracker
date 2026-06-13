"""API-эндпоинты для управления установщиками."""

from __future__ import annotations

import csv
import io
from typing import Annotated, Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.auth_middleware import get_current_user
from app.database import get_db
from app import models, schemas

router = APIRouter(prefix="/installers", tags=["installers"])

DbSession = Annotated[Session, Depends(get_db)]
CurrentUser = Annotated[dict[str, Any], Depends(get_current_user)]


class ImportResultRow(BaseModel):
    row: int
    name: str
    phone: str
    status: str  # "created", "skipped_duplicate", "error"
    message: Optional[str] = None


class ImportResultSummary(BaseModel):
    total: int
    created: int
    skipped: int
    errors: int
    details: List[ImportResultRow]


_HEADER_ALIASES: Dict[str, str] = {
    "name": "name",
    "фио": "name",
    "fio": "name",
    "имя": "name",
    "phone": "phone",
    "телефон": "phone",
    "tel": "phone",
}


def _normalize_header(header: str) -> str:
    """Приводит заголовок колонки к стандартному ключу (name / phone)."""
    return _HEADER_ALIASES.get(header.strip().lower(), header.strip().lower())


def _parse_csv(content: bytes) -> List[Dict[str, str]]:
    """Парсит CSV-файл; ожидаются колонки name/ФИО и phone/Телефон."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, str]] = []
    for row in reader:
        normalized: Dict[str, str] = {}
        for key, value in row.items():
            if key:
                normalized[_normalize_header(key)] = (value or "").strip()
        rows.append(normalized)
    return rows


def _parse_xlsx(content: bytes) -> List[Dict[str, str]]:
    """Парсит XLSX-файл; ожидаются колонки name/ФИО и phone/Телефон."""
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    headers: List[str] = []
    rows: List[Dict[str, str]] = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index == 0:
            headers = [_normalize_header(str(cell or "")) for cell in row]
            continue
        row_data: Dict[str, str] = {}
        for col_index, cell in enumerate(row):
            if col_index < len(headers) and headers[col_index]:
                row_data[headers[col_index]] = (
                    str(cell).strip() if cell is not None else ""
                )
        rows.append(row_data)

    workbook.close()
    return rows


def _get_installer_or_404(db: Session, installer_id: int) -> models.Installer:
    """Возвращает установщика по ID или 404."""
    installer = (
        db.query(models.Installer)
        .filter(models.Installer.id == installer_id)
        .first()
    )
    if installer is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Установщик не найден",
        )
    return installer


def _check_installer_ownership(
    installer: models.Installer,
    current_user: dict[str, Any],
) -> None:
    """Проверяет, что установщик принадлежит текущему пользователю (кроме admin)."""
    if current_user["role"] == "admin":
        return
    if installer.manager_id != current_user["manager_id"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Нет доступа к этому установщику",
        )


@router.get("", response_model=List[schemas.InstallerOut])
def list_installers(db: DbSession, current_user: CurrentUser) -> List[models.Installer]:
    """Список установщиков: менеджер — только свои, администратор — все."""
    query = db.query(models.Installer)

    if current_user["role"] == "manager":
        query = query.filter(
            models.Installer.manager_id == current_user["manager_id"]
        )

    return query.order_by(models.Installer.name).all()


@router.post(
    "",
    response_model=schemas.InstallerOut,
    status_code=status.HTTP_201_CREATED,
)
def create_installer(
    installer_data: schemas.InstallerCreate,
    db: DbSession,
    current_user: CurrentUser,
) -> models.Installer:
    """Создать установщика, привязанного к текущему пользователю."""
    existing = (
        db.query(models.Installer)
        .filter(models.Installer.phone == installer_data.phone)
        .first()
    )
    if existing is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Установщик с таким телефоном уже существует в системе",
        )

    installer = models.Installer(
        name=installer_data.name,
        phone=installer_data.phone,
        manager_id=current_user["manager_id"],
        is_active=True,
    )
    db.add(installer)
    db.commit()
    db.refresh(installer)
    return installer


@router.post("/import", response_model=ImportResultSummary)
async def import_installers(
    db: DbSession,
    current_user: CurrentUser,
    file: UploadFile = File(...),
) -> ImportResultSummary:
    """Импорт установщиков из CSV или XLSX. Привязка к текущему менеджеру."""
    filename = (file.filename or "").lower()
    content = await file.read()

    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживаются только файлы .csv и .xlsx",
        )

    details: List[ImportResultRow] = []
    created = 0
    skipped = 0
    errors = 0
    seen_phones: set[str] = set()
    manager_id = current_user["manager_id"]

    for idx, row_data in enumerate(rows, start=1):
        name = (row_data.get("name") or "").strip()
        phone = (row_data.get("phone") or "").strip()

        if not name or not phone:
            errors += 1
            details.append(
                ImportResultRow(
                    row=idx,
                    name=name,
                    phone=phone,
                    status="error",
                    message="Не заполнены ФИО или телефон",
                )
            )
            continue

        if phone in seen_phones:
            skipped += 1
            details.append(
                ImportResultRow(
                    row=idx,
                    name=name,
                    phone=phone,
                    status="skipped_duplicate",
                    message="Дубликат телефона в файле",
                )
            )
            continue

        existing = (
            db.query(models.Installer)
            .filter(models.Installer.phone == phone)
            .first()
        )
        if existing is not None:
            skipped += 1
            if existing.manager_id == manager_id:
                msg = "Установщик с таким телефоном уже существует у вас"
            else:
                msg = "Телефон уже используется другим менеджером"
            details.append(
                ImportResultRow(
                    row=idx,
                    name=name,
                    phone=phone,
                    status="skipped_duplicate",
                    message=msg,
                )
            )
            continue

        installer = models.Installer(
            name=name,
            phone=phone,
            manager_id=manager_id,
            is_active=True,
        )
        db.add(installer)
        seen_phones.add(phone)
        created += 1
        details.append(
            ImportResultRow(
                row=idx,
                name=name,
                phone=phone,
                status="created",
            )
        )

    db.commit()

    return ImportResultSummary(
        total=len(rows),
        created=created,
        skipped=skipped,
        errors=errors,
        details=details,
    )


@router.patch("/{installer_id}", response_model=schemas.InstallerOut)
def update_installer(
    installer_id: int,
    installer_data: schemas.InstallerUpdate,
    db: DbSession,
    current_user: CurrentUser,
) -> models.Installer:
    """Обновить данные установщика (только своего)."""
    installer = _get_installer_or_404(db, installer_id)
    _check_installer_ownership(installer, current_user)

    if installer_data.name is not None:
        installer.name = installer_data.name

    if installer_data.phone is not None:
        existing = (
            db.query(models.Installer)
            .filter(
                models.Installer.phone == installer_data.phone,
                models.Installer.id != installer.id,
            )
            .first()
        )
        if existing is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Установщик с таким телефоном уже существует в системе",
            )
        installer.phone = installer_data.phone

    if installer_data.is_active is not None:
        installer.is_active = installer_data.is_active

    db.commit()
    db.refresh(installer)
    return installer


@router.delete("/{installer_id}", response_model=schemas.InstallerOut)
def deactivate_installer(
    installer_id: int,
    db: DbSession,
    current_user: CurrentUser,
) -> models.Installer:
    """Деактивировать установщика (мягкое удаление, только своего)."""
    installer = _get_installer_or_404(db, installer_id)
    _check_installer_ownership(installer, current_user)

    installer.is_active = False
    db.commit()
    db.refresh(installer)
    return installer
